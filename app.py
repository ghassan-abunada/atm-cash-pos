import base64
import io
import os
from datetime import datetime, timezone

import xlrd
import openpyxl
from flask import Flask, render_template, request, send_file, jsonify
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB max

# ── Supabase client (optional — app still works without it) ──
_SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
_SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '')

try:
    from supabase import create_client
    _db = create_client(_SUPABASE_URL, _SUPABASE_KEY) if _SUPABASE_URL and _SUPABASE_KEY else None
except Exception:
    _db = None


def _require_db():
    if not _db:
        return None, jsonify({'error': 'Database not configured. Set SUPABASE_URL and SUPABASE_KEY.'}), 503
    return _db, None, None


# ── Report parsing ────────────────────────────────────────────

def parse_cassette_balances(file_bytes):
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheets()[0]
    datemode = wb.datemode

    terminals = {}
    for row_idx in range(4, ws.nrows):
        row = ws.row_values(row_idx)
        terminal_id = str(row[0]).strip()

        if not terminal_id:
            continue
        if 'Total' in terminal_id or 'Outstanding' in terminal_id:
            continue

        est_cash_out = None
        if isinstance(row[7], float) and row[7] > 40000:
            try:
                est_cash_out = xlrd.xldate_as_datetime(row[7], datemode)
            except Exception:
                est_cash_out = None

        terminals[terminal_id] = {
            'terminal_id': terminal_id,
            'site_name': str(row[2]).strip(),
            'cash_balance': row[4] if isinstance(row[4], (int, float)) else 0,
            'est_cash_out': est_cash_out,
            'last_error': str(row[8]).strip() if row[8] else '',
            'last_communication': str(row[9]).strip() if row[9] else '',
            'last_cash_wd': str(row[10]).strip() if row[10] else '',
            '_row_order': row_idx,
        }

    return terminals


def parse_cash_projection(file_bytes):
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheets()[0]

    projections = {}
    for row_idx in range(3, ws.nrows):
        row = ws.row_values(row_idx)
        terminal_id = str(row[0]).strip()

        if not terminal_id:
            continue

        projections[terminal_id] = {
            'days_until_load': int(row[6]) if isinstance(row[6], (int, float)) else None,
            'suggested_cash_add': int(row[8]) if isinstance(row[8], (int, float)) else None,
        }

    return projections


def get_city_lookup(wb):
    ws = wb['City']
    city_lookup = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1]:
            terminal_id = str(row[1]).strip()
            city = str(row[3]).strip() if row[3] else ''
            city_lookup[terminal_id] = city
    return city_lookup


def is_tulsa(city):
    return bool(city) and city.lower() in ('tulsa', 'tulsa_jd')


# ── Excel template writing ────────────────────────────────────

def clear_data_rows(ws, start_row, num_cols):
    if ws.max_row >= start_row:
        for row_idx in range(start_row, ws.max_row + 1):
            for col_idx in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=col_idx).value = None


def write_sheet_rows(ws, rows, col_keys, start_row=3):
    clear_data_rows(ws, start_row, len(col_keys))
    for i, row_data in enumerate(rows):
        row_num = start_row + i
        for col_offset, key in enumerate(col_keys, start=1):
            ws.cell(row=row_num, column=col_offset).value = row_data.get(key, '')


def update_template(template_bytes, cassette_data, projection_data):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=True)

    city_lookup = get_city_lookup(wb)

    sorted_terminals = sorted(cassette_data.values(), key=lambda x: x['_row_order'])

    all_rows = []
    for t in sorted_terminals:
        tid = t['terminal_id']
        proj = projection_data.get(tid, {})
        city = city_lookup.get(tid, '')

        all_rows.append({
            'terminal_id': tid,
            'site_name': t['site_name'],
            'city': city,
            'days_until_load': proj.get('days_until_load', ''),
            'suggested_cash_add': proj.get('suggested_cash_add', ''),
            'cash_remaining': t['cash_balance'],
            'last_transaction': t['last_cash_wd'],
            'est_cash_out': t['est_cash_out'],
            'last_error': t['last_error'],
            'last_communication': t['last_communication'],
        })

    cash_pos_cols = [
        'terminal_id', 'site_name', 'city',
        'days_until_load', 'suggested_cash_add',
        'cash_remaining', 'last_transaction',
    ]
    write_sheet_rows(wb['CashPosition'], all_rows, cash_pos_cols)

    tulsa_rows = [r for r in all_rows if is_tulsa(r['city'])]
    write_sheet_rows(wb['Tulsa'], tulsa_rows, cash_pos_cols)

    error_rep_cols = [
        'terminal_id', 'site_name', 'cash_remaining',
        'est_cash_out', 'last_error',
        'last_communication', 'last_transaction',
    ]
    write_sheet_rows(wb['Error Rep.'], all_rows, error_rep_cols)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, all_rows, tulsa_rows


# ── PDF generation ────────────────────────────────────────────

class ATMDriverPDF(FPDF):
    def footer(self):
        self.set_y(-12)
        self.set_font('Helvetica', 'I', 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 6, f'ATM Cash Routes  |  Page {self.page_no()}', align='C')


def build_pdf(drivers_data, report_date):
    pdf = ATMDriverPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.set_margins(15, 15, 15)

    COL_W = [28, 62, 26, 18, 24, 22]  # total 180 mm = A4 210 - 15mm margins each side

    def draw_table_header(pdf):
        headers = ['Terminal ID', 'Site Name', 'City', 'Days Left', 'Remaining', 'Cash to Add']
        pdf.set_fill_color(26, 43, 74)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 8)
        for w, h in zip(COL_W, headers):
            pdf.cell(w, 8, h, border=1, fill=True)
        pdf.ln(8)

    for driver in drivers_data:
        pdf.add_page()
        pdf.set_font('Helvetica', 'B', 15)
        pdf.set_text_color(26, 43, 74)
        pdf.cell(0, 10, f"Driver: {driver['name']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 6, f"Date: {report_date}  |  {len(driver['terminals'])} terminals",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(4)

        draw_table_header(pdf)

        driver_total = 0
        for j, t in enumerate(driver['terminals']):
            fill = j % 2 == 0
            if fill:
                pdf.set_fill_color(247, 250, 255)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font('Helvetica', '', 8)

            cash_to_add = t.get('cash_to_add') or 0
            driver_total += cash_to_add
            days = t.get('days_until_load')
            days_str = str(int(days)) if days is not None else 'OVERDUE'

            site = t.get('site_name', '')
            if len(site) > 37:
                site = site[:36] + '...'

            city = (t.get('city') or '').title()
            if len(city) > 16:
                city = city[:15] + '...'

            row_vals = [
                t.get('terminal_id', ''),
                site,
                city,
                days_str,
                f"${t.get('cash_remaining', 0):,.0f}",
                f"${cash_to_add:,.0f}",
            ]
            for w, v in zip(COL_W, row_vals):
                pdf.cell(w, 7, v, border=1, fill=fill)
            pdf.ln(7)

        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(220, 230, 245)
        pdf.set_text_color(0, 0, 0)
        label_w = sum(COL_W[:-1])
        pdf.cell(label_w, 8, f"  Total -- {len(driver['terminals'])} terminals",
                 border=1, fill=True)
        pdf.cell(COL_W[-1], 8, f"${driver_total:,.0f}", border=1, fill=True)
        pdf.ln(8)

    # Summary page
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(26, 43, 74)
    pdf.cell(0, 10, 'Route Summary -- All Drivers', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, report_date, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)

    SUM_W = [90, 35, 55]  # total 180 mm
    sum_headers = ['Driver', 'Terminals', 'Total Cash to Load']
    pdf.set_fill_color(26, 43, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 9)
    for w, h in zip(SUM_W, sum_headers):
        pdf.cell(w, 8, h, border=1, fill=True)
    pdf.ln(8)

    grand_terminals = 0
    grand_cash = 0
    for j, driver in enumerate(drivers_data):
        fill = j % 2 == 0
        pdf.set_fill_color(247, 250, 255) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 9)
        drv_cash = sum((t.get('cash_to_add') or 0) for t in driver['terminals'])
        drv_count = len(driver['terminals'])
        grand_terminals += drv_count
        grand_cash += drv_cash
        pdf.cell(SUM_W[0], 7, driver['name'], border=1, fill=fill)
        pdf.cell(SUM_W[1], 7, str(drv_count), border=1, fill=fill)
        pdf.cell(SUM_W[2], 7, f"${drv_cash:,.0f}", border=1, fill=fill)
        pdf.ln(7)

    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(26, 43, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(SUM_W[0], 9, 'GRAND TOTAL', border=1, fill=True)
    pdf.cell(SUM_W[1], 9, str(grand_terminals), border=1, fill=True)
    pdf.cell(SUM_W[2], 9, f"${grand_cash:,.0f}", border=1, fill=True)
    pdf.ln(9)

    return bytes(pdf.output())


# ── Routes ────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/step2')
def step2():
    return render_template('step2.html')


@app.route('/step3')
def step3():
    return render_template('step3.html')


@app.route('/step4')
def step4():
    return render_template('step4.html')


@app.route('/process', methods=['POST'])
def process():
    try:
        cassette_file = request.files.get('cassette')
        projection_file = request.files.get('projection')
        template_file = request.files.get('template')

        missing = []
        if not cassette_file or not cassette_file.filename:
            missing.append('Cassette Balances report')
        if not projection_file or not projection_file.filename:
            missing.append('Cash Projection report')
        if not template_file or not template_file.filename:
            missing.append('Cash POS template')

        if missing:
            return jsonify({'error': f"Missing files: {', '.join(missing)}"}), 400

        cassette_data = parse_cassette_balances(cassette_file.read())
        projection_data = parse_cash_projection(projection_file.read())
        template_bytes = template_file.read()

        output, all_rows, tulsa_rows = update_template(
            template_bytes, cassette_data, projection_data
        )

        today = datetime.now().strftime('%m-%d-%Y')
        filename = f'Cash_Pos_{today}.xlsm'

        excel_b64 = base64.b64encode(output.getvalue()).decode('utf-8')

        terminals_list = [
            {
                'terminal_id': r['terminal_id'],
                'site_name': r['site_name'],
                'city': r['city'],
                'city_key': r['city'].strip().lower() if r['city'] else '',
                'days_until_load': r['days_until_load'] if r['days_until_load'] != '' else None,
                'suggested_cash_add': r['suggested_cash_add'] if r['suggested_cash_add'] != '' else None,
                'cash_remaining': r['cash_remaining'],
                'last_cash_wd': r['last_transaction'],
            }
            for r in all_rows
        ]

        return jsonify({
            'terminals': terminals_list,
            'excel_b64': excel_b64,
            'filename': filename,
            'total': len(all_rows),
            'tulsa': len(tulsa_rows),
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Driver API ────────────────────────────────────────────────

@app.route('/api/drivers', methods=['GET'])
def get_drivers():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    result = db.table('drivers').select('id, name').order('name').execute()
    return jsonify(result.data)


@app.route('/api/drivers', methods=['POST'])
def create_driver():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    try:
        result = db.table('drivers').insert({'name': name}).execute()
        return jsonify(result.data[0]), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/drivers/<driver_id>', methods=['DELETE'])
def delete_driver(driver_id):
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    db.table('drivers').delete().eq('id', driver_id).execute()
    return '', 204


# ── City assignment API ───────────────────────────────────────

@app.route('/api/city-assignments', methods=['GET'])
def get_city_assignments():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    result = db.table('city_assignments').select('city, driver_id').execute()
    return jsonify(result.data)


@app.route('/api/city-assignments', methods=['POST'])
def save_city_assignments():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    data = request.get_json()
    if not isinstance(data, list):
        return jsonify({'error': 'Expected array'}), 400

    now = datetime.now(timezone.utc).isoformat()
    records = []
    for item in data:
        city_key = (item.get('city') or '').strip().lower()
        if city_key:
            records.append({
                'city': city_key,
                'driver_id': item.get('driver_id') or None,
                'updated_at': now,
            })

    if records:
        db.table('city_assignments').upsert(records, on_conflict='city').execute()

    return jsonify({'saved': len(records)})


# ── Terminal assignment API ───────────────────────────────────

@app.route('/api/terminal-assignments', methods=['GET'])
def get_terminal_assignments():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    result = db.table('terminal_assignments').select('terminal_id, driver_id').execute()
    return jsonify(result.data)


@app.route('/api/terminal-assignments', methods=['POST'])
def save_terminal_assignments():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    data = request.get_json()
    if not isinstance(data, list):
        return jsonify({'error': 'Expected array'}), 400
    now = datetime.now(timezone.utc).isoformat()
    records = [
        {'terminal_id': r['terminal_id'], 'driver_id': r.get('driver_id') or None, 'updated_at': now}
        for r in data if r.get('terminal_id')
    ]
    if records:
        db.table('terminal_assignments').upsert(records, on_conflict='terminal_id').execute()
    return jsonify({'saved': len(records)})


# ── Terminal status API ───────────────────────────────────────

@app.route('/api/terminal-status', methods=['GET'])
def get_terminal_status():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    result = db.table('terminal_status').select('terminal_id, status').execute()
    return jsonify(result.data)


@app.route('/api/terminal-status', methods=['POST'])
def save_terminal_status():
    db, err_resp, code = _require_db()
    if err_resp:
        return err_resp, code
    data = request.get_json()
    if not isinstance(data, list):
        return jsonify({'error': 'Expected array'}), 400
    now = datetime.now(timezone.utc).isoformat()
    records = [
        {'terminal_id': r['terminal_id'], 'status': r['status'], 'updated_at': now}
        for r in data if r.get('terminal_id') and r.get('status')
    ]
    if records:
        db.table('terminal_status').upsert(records, on_conflict='terminal_id').execute()
    return jsonify({'saved': len(records)})


# ── PDF route ─────────────────────────────────────────────────

@app.route('/generate-pdf', methods=['POST'])
def generate_pdf():
    try:
        data = request.get_json() or {}
        drivers_data = data.get('drivers', [])
        if not drivers_data:
            return jsonify({'error': 'No driver data provided'}), 400

        report_date = datetime.now().strftime('%B %d, %Y')
        pdf_bytes = build_pdf(drivers_data, report_date)

        today = datetime.now().strftime('%m-%d-%Y')
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'ATM_Driver_Routes_{today}.pdf',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Startup ───────────────────────────────────────────────────

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f'\n  ATM Cash POS Generator')
    print(f'  Open http://localhost:{port} in your browser\n')
    app.run(host='0.0.0.0', port=port, debug=False)
